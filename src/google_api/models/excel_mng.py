import pandas as pd

from typing import Literal
from openpyxl import load_workbook


class ExcelMng(object):
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = load_workbook(filename=file_name, read_only=True)

    def show_row_count(self, col_name: str) -> (int or Literal[False]):
        """指定項目のカウント数を返す.

        params:
          col_name(str): target Column Name
        return:
          Result count(int) or False
        """
        for sheet in self.wb.sheetnames:
            data = list(self.wb[sheet].values)
            # Define one row name as a column name
            df = pd.DataFrame(data[1:], columns=data[0])

            if col_name in list(df.columns):
                return df[col_name].count()
            return False
